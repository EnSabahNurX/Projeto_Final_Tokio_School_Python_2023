import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv
from io import StringIO, BytesIO
from datetime import datetime, date, timedelta
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    Response,
    make_response,
)
from models import db, Veiculo, VehicleType, Cliente, Reservation, Categoria
from werkzeug.utils import secure_filename
from app import app


def check_maintenance_status():
    """
    Verifica o status de manutenção dos veículos e atualiza quando necessário.
    """
    vehicles_in_maintenance = Veiculo.query.filter_by(in_maintenance=True).all()
    today = date.today()

    for vehicle in vehicles_in_maintenance:
        if vehicle.next_maintenance_date and vehicle.next_maintenance_date <= today:
            vehicle.status = True
            vehicle.in_maintenance = False
            vehicle.next_maintenance_date = vehicle.last_maintenance_date + timedelta(
                days=180
            )
            db.session.commit()


def register_usage(vehicle):
    """
    Registra uma nova utilização do veículo e verifica a próxima manutenção.
    """
    vehicle.num_uses += 1

    if vehicle.num_uses >= vehicle.max_uses_before_maintenance:
        days_since_last_maintenance = (
            datetime.now().date() - vehicle.last_maintenance_date
        ).days
        if days_since_last_maintenance >= 30:
            vehicle.next_maintenance_date = vehicle.last_maintenance_date + timedelta(
                days=30
            )

    db.session.commit()


def check_admin_session():
    """
    Verifica se a sessão de administrador está ativa para rotas que requerem autenticação de administrador.
    Redireciona para a página de login se não houver uma sessão de administrador ativa.
    """
    # Lista de rotas que requerem autenticação de administrador
    admin_routes = ["/admin", "/add_vehicle", "/edit_vehicle/", "/delete_vehicle/"]

    if request.path in admin_routes:
        if "admin" not in session:
            return redirect(url_for("login"))


def login():
    """
    Rota de login para administradores.
    Verifica se já há uma sessão de administrador ativa ou tenta autenticar um administrador.
    """
    # Verifica se já há uma sessão de admin ativa
    if "admin" in session:
        return redirect(url_for("admin_panel"))

    # Administradores temporários (em um projeto real, deve ser armazenado de forma segura)
    admins = {"admin": "password"}

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username in admins and admins[username] == password:
            # Definir a sessão de administrador como ativa
            session["admin"] = True
            return redirect(url_for("admin_panel"))

    return render_template("login.html")


def admin_panel():
    """
    Rota para o painel de administração.
    Exibe informações sobre veículos, alertas de manutenção, status de estoque e mais.
    """
    # Chamar a função para verificar a manutenção do veículo
    check_maintenance_status()

    # Consultar todos os veículos no banco de dados
    vehicles = Veiculo.query.all()
    # Lógica para verificar veículos que precisam de manutenção
    today = date.today()
    vehicles_needing_maintenance = [
        vehicle
        for vehicle in vehicles
        if vehicle.next_maintenance_date
        and vehicle.next_maintenance_date <= (today + timedelta(days=30))
        and not vehicle.in_maintenance
    ]
    if vehicles_needing_maintenance:
        # Montar a mensagem de alerta
        alert_message = "Atenção: Os seguintes veículos precisam de manutenção:\n"
        for vehicle in vehicles_needing_maintenance:
            alert_message += f"{vehicle.brand} {vehicle.model} ({vehicle.type.value})\n"

        # Enviar a mensagem de alerta para a página usando a função flash
        flash(alert_message, "warning")

    # Lógica para verificar veículos que precisam de legalização
    vehicles_needing_legalization = [
        vehicle
        for vehicle in vehicles
        if vehicle.next_legalization_date
        and (vehicle.next_legalization_date - today).days <= 30
        and not vehicle.in_maintenance
    ]
    if vehicles_needing_legalization:
        # Montar a mensagem de alerta
        alert_message = "Atenção: Os seguintes veículos precisam de legalização:\n"
        for vehicle in vehicles_needing_legalization:
            alert_message += f'{vehicle.brand} {vehicle.model} ({vehicle.type.value}) - Próxima Legalização: {vehicle.next_legalization_date.strftime("%d/%m/%Y")}\n'

        # Enviar a mensagem de alerta para a página usando a função flash
        flash(alert_message, "warning")

    # Verificar se a data de próxima legalização é None e, se for, atribuir uma data futura
    for vehicle in vehicles:
        if not vehicle.next_legalization_date:
            vehicle.next_legalization_date = date.today() + timedelta(days=365)

    # Verificar se há veículos suficientes no estoque
    num_veiculos = len(vehicles)
    num_clientes = len(Cliente.query.all())
    estoque_suficiente = num_veiculos >= num_clientes + 5

    # sempre que a página inicial for carregada, ela verificará se há veículos suficientes no estoque e exibirá uma mensagem de aviso caso contrário.
    if not estoque_suficiente:
        flash(
            "Atenção: O estoque de veículos está baixo. Considere adicionar mais veículos para atender à demanda.",
            "warning",
        )

    # Passar a variável 'date' para o template 'admin.html'
    return render_template(
        "admin.html",
        vehicles=vehicles,
        date=today,
        estoque_suficiente=estoque_suficiente,
    )


def list_vehicles():
    """
    Lista todos os veículos disponíveis e verifica o estoque suficiente.
    """
    vehicles = Veiculo.query.all()
    today = date.today()

    # Verificar se há veículos suficientes no estoque
    num_veiculos = len(vehicles)
    num_clientes = len(Cliente.query.all())
    estoque_suficiente = num_veiculos >= num_clientes + 5

    return render_template(
        "list_vehicles.html",
        vehicles=vehicles,
        date=today,
        estoque_suficiente=estoque_suficiente,
    )


def view_vehicle(id):
    """
    Exibe detalhes de um veículo específico com base no ID fornecido.
    """
    vehicle = Veiculo.query.get_or_404(id)
    imagens_paths = vehicle.imagens.split(",") if vehicle.imagens else []

    return render_template(
        "view_vehicle.html",
        vehicle=vehicle,
        imagens_paths=imagens_paths,
    )


def add_vehicle():
    """
    Adiciona um novo veículo ao banco de dados a partir dos dados do formulário.
    """
    if request.method == "POST":
        # Obter dados do formulário
        type = request.form["vehicle-type"]
        brand = request.form["brand"]
        model = request.form["model"]
        year = int(request.form["year"])
        price_per_day = float(request.form["price_per_day"])

        # Determinar a categoria do veículo
        categoria_nome = "Silver"
        if price_per_day > 250:
            categoria_nome = "Gold"
        elif price_per_day <= 50:
            categoria_nome = "Econômico"

        # Obter a categoria do banco de dados
        categoria = Categoria.query.filter_by(nome=categoria_nome).first()

        # Processar o upload das imagens
        imagens = request.files.getlist("imagens")
        imagens_paths = []
        for imagem in imagens:
            filename = secure_filename(imagem.filename)
            path = filename  # Caminho relativo à pasta 'static'
            full_path = os.path.join(app.config["UPLOAD_FOLDER"], path)
            imagem.save(full_path)
            imagens_paths.append(path)

        # Criar um novo objeto Veiculo e adicioná-lo ao banco de dados
        novo_veiculo = Veiculo(
            type=VehicleType[type.upper()],
            brand=brand,
            model=model,
            year=year,
            price_per_day=price_per_day,
            categoria=categoria,
        )

        # Salvar os caminhos das imagens
        novo_veiculo.imagens = ",".join(imagens_paths)

        # Validar o ano
        if not isinstance(int(request.form["year"]), int):
            flash("O ano deve ser um número inteiro.", category="danger")
            return False

        # Validar a diária
        if float(request.form["price_per_day"]) <= 0:
            flash("A diária deve ser um número positivo.", category="danger")
            return False

        # Validar as datas de legalização
        if (
            request.form["last_legalization_date"]
            >= request.form["next_legalization_date"]
        ):
            flash(
                "A última data de legalização deve ser anterior à próxima data de legalização.",
                category="danger",
            )
            return False

        # Salvar o veículo
        novo_veiculo.initialize_vehicle()
        db.session.add(novo_veiculo)
        db.session.commit()  # Salvar o novo veículo no banco de dados

        # Redirecionar para o painel de administração com mensagem de sucesso
        flash("Novo veículo adicionado com sucesso!", "success")
        return redirect(url_for("admin_panel"))

    return render_template("add_vehicle.html")


def edit_vehicle(id):
    """
    Permite a edição dos detalhes de um veículo existente com base no ID fornecido.
    """
    # Obter o veículo pelo ID
    vehicle = Veiculo.query.get_or_404(id)

    if request.method == "POST":
        # Obter os dados do formulário
        type = request.form["vehicle_type"]
        brand = request.form["brand"]
        model = request.form["model"]
        year = int(request.form["year"])
        price_per_day = float(request.form["price_per_day"])
        last_maintenance_date_str = request.form["last_maintenance_date"]
        next_maintenance_date_str = request.form["next_maintenance_date"]
        last_legalization_date_str = request.form["last_legalization_date"]
        next_legalization_date_str = request.form["next_legalization_date"]
        available_from_str = request.form["available_from"]
        max_uses_before_maintenance = int(request.form["max_uses_before_maintenance"])

        # Determinar a categoria do veículo
        categoria_nome = "Silver"
        if price_per_day > 250:
            categoria_nome = "Gold"
        elif price_per_day <= 50:
            categoria_nome = "Econômico"

        # Obter a categoria do banco de dados
        categoria = Categoria.query.filter_by(nome=categoria_nome).first()

        # Converter as datas do formulário em objetos date
        vehicle.last_maintenance_date = datetime.strptime(
            last_maintenance_date_str, "%Y-%m-%d"
        ).date()
        vehicle.next_maintenance_date = datetime.strptime(
            next_maintenance_date_str, "%Y-%m-%d"
        ).date()
        vehicle.last_legalization_date = datetime.strptime(
            last_legalization_date_str, "%Y-%m-%d"
        ).date()
        vehicle.next_legalization_date = datetime.strptime(
            next_legalization_date_str, "%Y-%m-%d"
        ).date()
        vehicle.available_from = datetime.strptime(
            available_from_str, "%Y-%m-%d"
        ).date()

        # Validar o ano
        if not isinstance(int(request.form["year"]), int):
            flash("O ano deve ser um número inteiro.", category="danger")
            return False

        # Validar a diária
        if float(request.form["price_per_day"]) <= 0:
            flash("A diária deve ser um número positivo.", category="danger")
            return False

        # Validar as datas de legalização
        if (
            request.form["last_legalization_date"]
            >= request.form["next_legalization_date"]
        ):
            flash(
                "A última data de legalização deve ser anterior à próxima data de legalização.",
                category="danger",
            )
            return False

        # Processar o upload das imagens
        imagens = request.files.getlist("imagens")
        imagens_paths = vehicle.imagens.split(",")

        for imagem in imagens:
            if imagem.filename == "":
                # A imagem está vazia, ignorá-la
                continue

            filename = secure_filename(imagem.filename)
            path = filename  # Caminho relativo à pasta 'static'
            full_path = os.path.join(app.config["UPLOAD_FOLDER"], path)
            imagem.save(full_path)
            imagens_paths.append(path)

        # Atualizar os dados do veículo com base no formulário
        vehicle.type = VehicleType[type.upper()]
        vehicle.brand = brand
        vehicle.model = model
        vehicle.year = year
        vehicle.price_per_day = price_per_day
        vehicle.categoria = categoria
        vehicle.max_uses_before_maintenance = max_uses_before_maintenance

        # Atualizar os caminhos das imagens
        vehicle.imagens = ",".join(imagens_paths)

        # Salvar as alterações no banco de dados
        db.session.commit()

        # Redirecionar para a visualização do veículo com mensagem de sucesso
        flash("Detalhes do veículo atualizados com sucesso!", "success")
        return redirect(url_for("view_vehicle", id=id))

    return render_template("edit_vehicle.html", vehicle=vehicle)


def delete_vehicle(id):
    """
    Remove um veículo do banco de dados com base no ID fornecido.
    """
    # Obter o veículo pelo ID
    vehicle = Veiculo.query.get_or_404(id)

    if request.method == "POST":
        # Remover o veículo do banco de dados
        db.session.delete(vehicle)
        db.session.commit()

        # Redirecionar para o painel de administração com mensagem de sucesso
        flash("Veículo removido com sucesso!", "success")
        return redirect(url_for("admin_panel"))

    return render_template("delete_vehicle.html", vehicle=vehicle)


def delete_image(image_path, vehicle_id):
    """
    Remove uma imagem associada a um veículo específico.
    """
    # Encontra o veículo no banco de dados pelo ID
    vehicle = Veiculo.query.get(vehicle_id)
    if not vehicle:
        flash("Veículo não encontrado.", "error")
        return redirect(url_for("edit_vehicle", id=vehicle_id))

    if request.method == "POST" or request.form.get("_method") == "DELETE":
        # Verifica se a imagem está associada ao veículo
        if image_path in vehicle.imagens.split(","):
            try:
                # Remove a imagem do servidor
                os.remove(os.path.join(app.config["UPLOAD_FOLDER"], image_path))
            except Exception as e:
                flash(
                    "Imagem não existe na base de dados, então foi atualizado a lista e removida a imagem não existente!",
                    "warning",
                )

            # Atualiza o registro do veículo no banco de dados para refletir a remoção da imagem
            imagens = vehicle.imagens.split(",")
            imagens.remove(image_path)
            vehicle.imagens = ",".join(imagens).lstrip("").lstrip(",")

            # Salva as alterações no banco de dados
            db.session.commit()

            flash("Imagem removida com sucesso!", "success")

        return redirect(url_for("edit_vehicle", id=vehicle_id))

    else:
        flash("Método de requisição inválido.", "error")
        return redirect(
            url_for(
                "edit_vehicle",
                id=vehicle_id,
            )
        )


def logout():
    """
    Encerra a sessão do administrador.
    """
    # Remove a sessão de administrador
    session.pop("admin", None)
    return redirect(url_for("login"))


def legalize_vehicle(id):
    """
    Legaliza um veículo atualizando as datas de legalização.
    """
    # Obtém o veículo pelo ID
    vehicle = Veiculo.query.get_or_404(id)

    # Atualiza a data da última legalização
    vehicle.last_legalization_date = date.today()

    # Atualiza a data da próxima legalização (1 ano após a última legalização)
    vehicle.next_legalization_date = vehicle.last_legalization_date + timedelta(
        days=365
    )

    # Salva o histórico de legalizações
    vehicle.legalization_history += (
        f'Legalização realizada {vehicle.last_legalization_date.strftime("%Y-%m-%d")};'
    )

    # Salva as alterações no banco de dados
    db.session.commit()

    flash("Veículo legalizado com sucesso!", "info")

    return redirect(url_for("admin_panel"))


def maintenance_vehicle(id):
    """
    Inicia ou conclui a manutenção de um veículo.
    """
    vehicle = Veiculo.query.get_or_404(id)

    if request.method == "POST":
        if "maintenance" in request.form:
            # Inicia a manutenção do veículo
            vehicle.status = False
            vehicle.in_maintenance = True
            vehicle.last_maintenance_date = date.today()
            vehicle.next_maintenance_date = date.today() + timedelta(days=6 * 30)
            vehicle.maintenance_history += f'Manutenção iniciada {vehicle.last_maintenance_date.strftime("%Y-%m-%d")};'

            db.session.commit()

            flash("Veículo enviado para manutenção com sucesso!", "success")

        elif "complete_maintenance" in request.form:
            # Conclui a manutenção do veículo
            vehicle.status = True
            vehicle.in_maintenance = False
            vehicle.next_maintenance_date = vehicle.last_maintenance_date + timedelta(
                days=180
            )
            vehicle.maintenance_history += f'Manutenção concluída {vehicle.last_maintenance_date.strftime("%Y-%m-%d")};'

            db.session.commit()

            flash("Veículo concluiu a manutenção com sucesso!", "success")

        return redirect(url_for("admin_panel"))

    return render_template(
        "maintenance_vehicle.html",
        vehicle=vehicle,
    )


def register_usage_route(vehicle_id):
    """
    Registra a utilização de um veículo pelo ID e redireciona para a página de edição do veículo.
    """
    # Obtém o veículo pelo ID
    vehicle = Veiculo.query.get_or_404(vehicle_id)

    # Registra a utilização e verifica a próxima manutenção
    register_usage(vehicle)

    # Redireciona de volta para a página de edição do veículo
    flash("Utilização registrada com sucesso!", "success")
    return redirect(
        url_for(
            "edit_vehicle",
            id=vehicle_id,
        )
    )


def categorias():
    """
    Exibe a lista de categorias e permite adicionar novas categorias.
    """
    if request.method == "POST":
        nome = request.form.get("nome")

        if not nome:
            flash("Por favor, forneça um nome para a categoria", "danger")
        else:
            categoria = Categoria(nome=nome)
            db.session.add(categoria)
            db.session.commit()
            flash("Categoria adicionada com sucesso", "success")

    categorias = Categoria.query.all()
    return render_template("admin_categorias.html", categorias=categorias)


def editar_categoria(id):
    """
    Permite editar uma categoria existente pelo seu ID.
    """
    categoria = Categoria.query.get(id)

    if request.method == "POST":
        nome = request.form.get("nome")

        if not nome:
            flash("Por favor, forneça um nome para a categoria", "danger")
        else:
            categoria.nome = nome
            db.session.commit()
            flash("Categoria editada com sucesso", "success")
            return redirect(url_for("categorias"))

    return render_template("editar_categoria.html", categoria=categoria)


def deletar_categoria(id):
    """
    Deleta uma categoria existente pelo seu ID.
    """
    categoria = Categoria.query.get(id)

    if categoria:
        db.session.delete(categoria)
        db.session.commit()
        flash("Categoria removida com sucesso", "success")

    return redirect(url_for("categorias"))


def list_clients():
    """
    Exibe a lista de clientes.
    """
    clients = Cliente.query.all()
    return render_template("list_clients.html", clients=clients)


def delete_client(id):
    """
    Deleta um cliente pelo seu ID.
    """
    cliente = Cliente.query.get_or_404(id)
    db.session.delete(cliente)
    db.session.commit()
    flash("Cliente excluído com sucesso!", "success")
    return redirect(url_for("list_clients"))


def admin_edit_client(id):
    """
    Permite a edição de informações de um cliente pelo seu ID.
    """
    cliente = Cliente.query.get_or_404(id)

    if request.method == "POST":
        # Atualiza os campos do cliente com base nos dados do formulário
        cliente.nome = request.form["nome"]
        cliente.apelido = request.form["apelido"]
        cliente.email = request.form["email"]
        cliente.telefone = request.form["telefone"]
        cliente.data_nascimento = datetime.strptime(
            request.form.get("data_nascimento"), "%Y-%m-%d"
        ).date()
        cliente.morada = request.form.get("morada")
        cliente.nif = request.form.get("nif")
        cliente.price_per_day = float(request.form["price_per_day"])
        if request.form.get("password"):
            cliente.password = request.form.get("password")

        # Define a categoria do cliente com base na diária escolhida
        if cliente.price_per_day > 250:
            categoria = "Gold"
        elif cliente.price_per_day <= 50:
            categoria = "Económico"
        else:
            categoria = "Silver"
        cliente.categoria = categoria

        db.session.commit()
        flash("Cliente atualizado com sucesso!", "success")
        return redirect(url_for("list_clients"))

    return render_template("admin_edit_client.html", cliente=cliente)


def export_csv():
    """
    Exporta dados dos veículos para um arquivo CSV.
    """
    # Consulta todos os veículos na tabela de Veículos
    vehicles = Veiculo.query.all()

    # Cria um objeto StringIO para armazenar o conteúdo CSV
    csv_buffer = StringIO()
    csv_writer = csv.writer(csv_buffer)

    # Escreve os títulos das colunas no arquivo CSV
    csv_writer.writerow(
        [
            "ID",
            "Tipo",
            "Marca",
            "Modelo",
            "Ano",
            "Diária (€)",
            "Categoria",
            "Status",
            "Em Manutenção",
            "Próxima Manutenção",
            "Próxima Legalização",
        ]
    )

    # Escreve os dados de cada veículo no arquivo CSV
    for vehicle in vehicles:
        csv_writer.writerow(
            [
                vehicle.id,
                vehicle.type.value,
                vehicle.brand,
                vehicle.model,
                vehicle.year,
                vehicle.price_per_day,
                vehicle.categoria.nome,
                "Disponível" if vehicle.status else "Indisponível",
                "Sim" if vehicle.in_maintenance else "Não",
                vehicle.next_maintenance_date.strftime("%d/%m/%Y")
                if vehicle.next_maintenance_date
                else "",
                vehicle.next_legalization_date.strftime("%d/%m/%Y")
                if vehicle.next_legalization_date
                else "",
            ]
        )

    # Prepare a resposta com o conteúdo CSV
    response = Response(csv_buffer.getvalue(), content_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=vehicles.csv"

    # Fecha o buffer
    csv_buffer.close()

    return response


def export_excel():
    """
    Exporta dados dos veículos para um arquivo Excel.
    """
    # Obtenha os dados dos veículos do banco de dados (substitua esta parte pelo seu código existente)
    vehicles = Veiculo.query.all()

    # Crie um DataFrame pandas com os dados dos veículos
    data = {
        "ID": [vehicle.id for vehicle in vehicles],
        "Tipo": [vehicle.type.value for vehicle in vehicles],
        "Marca": [vehicle.brand for vehicle in vehicles],
        "Modelo": [vehicle.model for vehicle in vehicles],
        "Ano": [vehicle.year for vehicle in vehicles],
        "Diária (€)": [vehicle.price_per_day for vehicle in vehicles],
        "Categoria": [vehicle.categoria.nome for vehicle in vehicles],
        "Status": [
            "Disponível" if vehicle.status else "Indisponível" for vehicle in vehicles
        ],
        "Em Manutenção": [
            "Sim" if vehicle.in_maintenance else "Não" for vehicle in vehicles
        ],
        "Próxima Manutenção": [
            vehicle.next_maintenance_date.strftime("%d/%m/%Y")
            if vehicle.next_maintenance_date
            else ""
            for vehicle in vehicles
        ],
        "Próxima Legalização": [
            vehicle.next_legalization_date.strftime("%d/%m/%Y")
            if vehicle.next_legalization_date
            else ""
            for vehicle in vehicles
        ],
    }
    df = pd.DataFrame(data)

    # Crie um objeto BytesIO para armazenar o arquivo Excel
    excel_output = BytesIO()

    # Use o Pandas para escrever o DataFrame no arquivo Excel
    df.to_excel(excel_output, index=False, engine="openpyxl", sheet_name="Veículos")

    # Carregue o arquivo Excel criado
    excel_output.seek(0)
    wb = Workbook()
    with BytesIO(excel_output.read()) as excel_read:
        wb = load_workbook(excel_read)

    # Acesse a planilha
    ws = wb.active

    # Estilize a planilha (você pode personalizar isso de acordo com suas necessidades)
    header_font = Font(bold=True, color="FFFFFFFF")  # Fonte para cabeçalho
    header_fill = PatternFill(
        start_color="333333", end_color="333333", fill_type="solid"
    )  # Cor de fundo para cabeçalho
    cell_alignment = Alignment(
        horizontal="center", vertical="center"
    )  # Alinhamento das células

    # Estilize o cabeçalho (primeira linha)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment

    # Estilize o restante das células
    for row in ws.iter_rows(min_row=2, max_row=len(vehicles) + 1):
        for cell in row:
            cell.alignment = cell_alignment

    # Salve o arquivo Excel atualizado
    excel_output = BytesIO()
    wb.save(excel_output)

    # Configurar a resposta HTTP para baixar o arquivo Excel
    excel_output.seek(0)
    response = make_response(excel_output.read())
    response.headers[
        "Content-Type"
    ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=veiculos.xlsx"

    return response
